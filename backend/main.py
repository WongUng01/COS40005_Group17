from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
import os
from supabaseClient import get_supabase_client
from fastapi import Query
from fastapi.responses import FileResponse
import shutil
import openpyxl
from fastapi import HTTPException
import time
from pydantic import BaseModel
from typing import List, Optional

app = FastAPI()

# CORS settings
origins = ["http://localhost:3000"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class StudentBase(BaseModel):
    student_id: str
    name: str
    graduation_status: Optional[bool] = False

class StudentCreate(StudentBase):
    pass

class Student(StudentBase):
    id: str

class StudentUnit(BaseModel):
    unit_code: str
    unit_name: str
    grade: Optional[str] = None
    completed: Optional[bool] = False

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_FOLDER), name="uploads")

@app.post("/upload_planner/")
async def upload_planner(file: UploadFile = File(...), year: str = Form(...)):
    year_folder = os.path.join(UPLOAD_FOLDER, year)
    os.makedirs(year_folder, exist_ok=True)

    file_location = os.path.join(year_folder, file.filename)
    with open(file_location, "wb") as f:
        f.write(await file.read())

    # Return relative URL for frontend
    url = f"/uploads/{year}/{file.filename}"
    return JSONResponse(content={"url": url})

# Endpoint to get uploaded PDFs
@app.get("/get_uploaded_pdfs/")
def get_uploaded_pdfs():
    # Organize files by year
    files_by_year = defaultdict(list)
    
    # List all files in the upload folder
    for filename in os.listdir(UPLOAD_FOLDER):
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        if os.path.isfile(file_path) and filename.endswith(".pdf"):
            # Extract the year from the file name (assuming it's in the format: '2024_filename.pdf')
            try:
                year = filename.split("_")[0]  # Extract year from the file name, assuming format is like '2024_filename.pdf'
                files_by_year[year].append(f"/uploads/{filename}")
            except Exception as e:
                print(f"Error extracting year from file name {filename}: {e}")
    
    # Convert defaultdict to a regular dict for returning as JSON
    return JSONResponse(content=dict(files_by_year))

@app.get("/list_planners/")
def list_planners(year: str = Query(...)):
    folder_path = os.path.join(UPLOAD_FOLDER, year)
    if not os.path.exists(folder_path):
        return []
    files = os.listdir(folder_path)
    return [f"/uploads/{year}/{file}" for file in files if file.endswith(".pdf")]

@app.delete("/delete_planner/")
async def delete_planner(year: str, filename: str):
    try:
        file_path = f"./uploads/{year}/{filename}"
        os.remove(file_path)
        return JSONResponse(content={"message": "File deleted successfully."})
    except FileNotFoundError:
        return JSONResponse(status_code=404, content={"message": "File not found."})
    except Exception as e:
        return JSONResponse(status_code=500, content={"message": str(e)})
    
@app.post("/upload_excel/")
async def upload_excel(files: list[UploadFile] = File(...), year: str = Form(...)):
    if not files:
        return JSONResponse(status_code=400, content={"error": "No files uploaded"})

    save_dir = f"excel_uploads/{year}"
    os.makedirs(save_dir, exist_ok=True)

    # Store data for each uploaded file
    all_data = []

    for file in files:
        if not file.filename.endswith((".xlsx", ".xls")):
            return JSONResponse(status_code=400, content={"error": "Only Excel files are allowed"})

        # Save the file with a unique filename using timestamp
        timestamp = int(time.time())
        filename = f"{timestamp}_{file.filename}"
        file_path = os.path.join(save_dir, filename)

        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        # Parse the uploaded file and store its data
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        data = [[cell.value for cell in row] for row in ws.iter_rows()]

        # Ensure we append an array (data for one file)
        all_data.append(data)

    return JSONResponse(content=all_data)

@app.get("/get_excel_data/{year}")
def get_excel_data(year: str):
    dir_path = f"excel_uploads/{year}"
    if not os.path.exists(dir_path):
        raise HTTPException(status_code=404, detail="No uploads for this year")

    files = os.listdir(dir_path)
    if not files:
        raise HTTPException(status_code=404, detail="No Excel file found")

    all_data = []

    for file_name in files:
        if file_name.endswith((".xlsx", ".xls")):
            file_path = os.path.join(dir_path, file_name)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Read data from the Excel file
            data = [[cell.value for cell in row] for row in ws.iter_rows()]
            all_data.append(data)

    if not all_data:
        raise HTTPException(status_code=404, detail="No valid Excel data found")

    return all_data

@app.get("/study_planners/")
def get_study_planners():
    supabase = get_supabase_client()
    response = supabase.table("hod_study_planner").select("*").execute()
    return response.data

@app.post("/study_planners/")
def create_study_planner(planner: dict):
    supabase = get_supabase_client()
    response = supabase.table("hod_study_planner").insert([planner]).execute()
    if response.error:
        raise HTTPException(status_code=500, detail=response.error.message)
    return response.data[0]

@app.get("/students")
def get_all_students():
    students_data = supabase.table("students").select("*").execute().data
    units_data = supabase.table("units").select("code", "is_required").execute().data
    student_units_data = supabase.table("student_units").select("*").execute().data

    required_units = {u['code'] for u in units_data if u['is_required']}

    # Construct graduation status
    student_graduation_status = []
    for student in students_data:
        sid = student["id"]
        units_taken = [u for u in student_units_data if u["student_id"] == sid and u["grade"] == "Pass"]
        passed_units = {u["unit_code"] for u in units_taken}
        is_graduated = required_units.issubset(passed_units)
        student_graduation_status.append({
            "id": sid,
            "name": student["name"],
            "graduated": is_graduated
        })

    return student_graduation_status

@app.get("/students/{student_id}")
def get_student_details(student_id: str):
    student = supabase.table("students").select("*").eq("id", student_id).single().execute().data
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    units_data = supabase.table("units").select("*").execute().data
    student_units_data = supabase.table("student_units").select("*").eq("student_id", student_id).execute().data

    unit_lookup = {u['code']: u for u in units_data}
    details = []

    for su in student_units_data:
        code = su['unit_code']
        unit = unit_lookup.get(code, {})
        details.append({
            "unit_code": code,
            "unit_name": unit.get("name", ""),
            "grade": su["grade"],
            "is_required": unit.get("is_required", False),
            "passed": su["grade"] == "Pass"
        })

    # Graduation logic
    required_units = {u['code'] for u in units_data if u['is_required']}
    passed_units = {su["unit_code"] for su in student_units_data if su["grade"] == "Pass"}
    graduated = required_units.issubset(passed_units)

    return {
        "id": student["id"],
        "name": student["name"],
        "graduated": graduated,
        "units": details
    }

# Test endpoint
@app.get("/ping")
def ping():
    return {"message": "Connected to FastAPI"}

@app.get("/test-connection")
def test_connection():
    try:
        client = get_supabase_client()
        if client:
            return {"message": "Connection to Supabase is successful!"}
        else:
            raise HTTPException(status_code=500, detail="Failed to initialize Supabase client.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error connecting to Supabase: {str(e)}")

# In-memory storage for units
units = []

@app.get("/units")
def get_units():
    return units

@app.post("/units")
def create_unit(unit: dict):
    unit["id"] = len(units) + 1
    units.append(unit)
    return unit

@app.put("/units/{unit_id}")
def update_unit(unit_id: int, updated: dict):
    for i in range(len(units)):
        if units[i]["id"] == unit_id:
            units[i].update(updated)
            return units[i]
    raise HTTPException(status_code=404, detail="Unit not found")

@app.delete("/units/{unit_id}")
def delete_unit(unit_id: int):
    global units
    units = [unit for unit in units if unit["id"] != unit_id]
    return {"message": "Deleted"}